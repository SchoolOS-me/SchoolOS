import csv
import io
import json

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils.text import slugify
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.constants import UserRole
from apps.accounts.permissions import is_school_admin, is_super_admin
from apps.academics.models import (
    AcademicYear,
    ClassTeacher,
    SchoolClass,
    Section,
    Student,
    Subject,
    Teacher,
    TeachingAssignment,
)
from apps.finances.services import subscriptions
from apps.finances.models import Price
from apps.multitenancy.constants import TenantType
from apps.multitenancy.models import Tenant
from .models import School
from .serializers import (
    SchoolAdminCreateSerializer,
    SchoolCreateSerializer,
    SchoolSerializer,
    SchoolSubscriptionAssignSerializer,
)

IMPORT_GROUP_FIELDS = {
    "classes": ["class_name", "class_code", "academic_year", "order"],
    "sections": ["class_code", "class_name", "section_name", "class_teacher_email"],
    "students": ["admission_no", "student_name", "class_code", "class_name", "section_name", "parent_phone"],
    "teachers": ["teacher_name", "email", "employee_id", "subjects", "class_teacher_of"],
}


def _current_academic_year_name():
    from datetime import date

    today = date.today()
    return f"{today.year}-{today.year + 1}"


def _current_academic_year_dates():
    from datetime import date

    today = date.today()
    return date(today.year, 1, 1), date(today.year, 12, 31)


def _normalize_key(value):
    return str(value or "").strip().lower()


def _make_import_password(prefix, index):
    return f"{prefix}{index:04d}!SchoolOS"


def _auth_bypass_enabled():
    permission_classes = settings.REST_FRAMEWORK.get("DEFAULT_PERMISSION_CLASSES", ())
    return "rest_framework.permissions.AllowAny" in permission_classes


def _is_super_admin_or_open(user):
    return _auth_bypass_enabled() or is_super_admin(user)


def _is_school_admin_for_school(user, school):
    return is_school_admin(user) and getattr(user, "school_id", None) == school.id


def _resolve_tenant_creator(request):
    if request.user and request.user.is_authenticated:
        return request.user

    user_model = get_user_model()
    creator = user_model.objects.filter(role=UserRole.SUPER_ADMIN).order_by("id").first()
    if creator:
        return creator

    creator = user_model.objects.order_by("id").first()
    if creator:
        return creator

    raw_seed = request.data.get("code") or request.data.get("name") or "school"
    seed = slugify(str(raw_seed)).replace("-", "")[:20] or "school"
    email = f"bootstrap.superadmin+{seed}@kcs.local"

    suffix = 1
    while user_model.objects.filter(email__iexact=email).exists():
        suffix += 1
        email = f"bootstrap.superadmin+{seed}{suffix}@kcs.local"

    bootstrap_user = user_model.objects.create_user(
        email=email,
        password=None,
        role=UserRole.SUPER_ADMIN,
        is_staff=True,
        is_superuser=True,
    )
    bootstrap_user.set_unusable_password()
    bootstrap_user.save(update_fields=["password"])
    return bootstrap_user


def _parse_uploaded_table(uploaded_file):
    file_name = (getattr(uploaded_file, "name", "") or "").lower()
    if file_name.endswith(".csv"):
      text = uploaded_file.read().decode("utf-8-sig")
      reader = csv.reader(io.StringIO(text))
      rows = [list(row) for row in reader if any(str(cell or "").strip() for cell in row)]
      if not rows:
          return [], []
      return [str(value).strip() for value in rows[0]], rows[1:]

    if file_name.endswith(".xlsx") or file_name.endswith(".xlsm") or file_name.endswith(".xltx"):
        from openpyxl import load_workbook

        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        all_rows = list(worksheet.iter_rows(values_only=True))
        rows = [
            ["" if value is None else str(value).strip() for value in row]
            for row in all_rows
            if any(str(value or "").strip() for value in row)
        ]
        if not rows:
            return [], []
        return rows[0], rows[1:]

    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def _rows_with_mapping(headers, rows, mapping):
    header_lookup = {str(header).strip(): index for index, header in enumerate(headers)}
    normalized_rows = []
    for row in rows:
        normalized = {}
        for target_field, source_header in mapping.items():
            if not source_header:
                continue
            source_index = header_lookup.get(str(source_header).strip())
            normalized[target_field] = str(row[source_index]).strip() if source_index is not None and source_index < len(row) else ""
        normalized_rows.append(normalized)
    return normalized_rows


def _run_school_bulk_import(school, payload):
    class_rows = payload.get("classes") or []
    section_rows = payload.get("sections") or []
    student_rows = payload.get("students") or []
    teacher_rows = payload.get("teachers") or []

    start_date, end_date = _current_academic_year_dates()
    academic_year = AcademicYear.objects.filter(tenant=school.tenant, is_active=True).first()
    if academic_year is None:
        academic_year = AcademicYear.objects.create(
            tenant=school.tenant,
            name=_current_academic_year_name(),
            start_date=start_date,
            end_date=end_date,
            is_active=True,
        )

    summary = {
        "classes": {"created": 0, "updated": 0, "skipped": 0},
        "sections": {"created": 0, "updated": 0, "skipped": 0},
        "students": {"created": 0, "updated": 0, "skipped": 0},
        "teachers": {"created": 0, "updated": 0, "skipped": 0},
    }
    errors = []

    class_code_map = {}
    class_name_map = {
        _normalize_key(item.name): item
        for item in SchoolClass.objects.filter(tenant=school.tenant, academic_year=academic_year)
    }
    section_map = {
        (_normalize_key(item.school_class.name), _normalize_key(item.name)): item
        for item in Section.objects.select_related("school_class").filter(tenant=school.tenant)
    }
    teacher_email_map = {
        _normalize_key(item.user.email): item
        for item in Teacher.objects.select_related("user").filter(tenant=school.tenant)
        if item.user and item.user.email
    }

    def add_error(group, index, detail):
        errors.append({"group": group, "row": index + 1, "detail": detail})

    def resolve_school_class(identifier):
        key = _normalize_key(identifier)
        return class_code_map.get(key) or class_name_map.get(key)

    def resolve_section(class_identifier, section_name):
        school_class = resolve_school_class(class_identifier)
        if school_class is None:
            return None, None
        return school_class, section_map.get((_normalize_key(school_class.name), _normalize_key(section_name)))

    for index, row in enumerate(class_rows):
        class_name = str(row.get("class_name") or "").strip()
        class_code = _normalize_key(row.get("class_code"))
        academic_year_name = str(row.get("academic_year") or "").strip()
        order_raw = str(row.get("order") or "").strip()

        if not class_name:
            add_error("classes", index, "class_name is required.")
            summary["classes"]["skipped"] += 1
            continue

        target_year = academic_year
        if academic_year_name and academic_year_name != academic_year.name:
            target_year = AcademicYear.objects.filter(tenant=school.tenant, name=academic_year_name).first()
            if target_year is None:
                target_year = AcademicYear.objects.create(
                    tenant=school.tenant,
                    name=academic_year_name,
                    start_date=start_date,
                    end_date=end_date,
                    is_active=False,
                )

        school_class = SchoolClass.objects.filter(
            tenant=school.tenant,
            academic_year=target_year,
            name__iexact=class_name,
        ).first()

        order = int(order_raw) if order_raw.isdigit() else index + 1
        if school_class is None:
            school_class = SchoolClass.objects.create(
                tenant=school.tenant,
                academic_year=target_year,
                name=class_name,
                order=order,
            )
            summary["classes"]["created"] += 1
        else:
            updates = []
            if school_class.order != order:
                school_class.order = order
                updates.append("order")
            if school_class.academic_year_id != target_year.id:
                school_class.academic_year = target_year
                updates.append("academic_year")
            if updates:
                school_class.save(update_fields=updates)
                summary["classes"]["updated"] += 1
            else:
                summary["classes"]["skipped"] += 1

        class_name_map[_normalize_key(class_name)] = school_class
        if class_code:
            class_code_map[class_code] = school_class

    pending_class_teacher_links = []

    for index, row in enumerate(section_rows):
        class_identifier = row.get("class_code") or row.get("class_name")
        section_name = str(row.get("section_name") or "").strip()
        class_teacher_email = str(row.get("class_teacher_email") or "").strip()

        if not class_identifier or not section_name:
            add_error("sections", index, "class_code and section_name are required.")
            summary["sections"]["skipped"] += 1
            continue

        school_class = resolve_school_class(class_identifier)
        if school_class is None:
            add_error("sections", index, f"Unknown class reference: {class_identifier}")
            summary["sections"]["skipped"] += 1
            continue

        section, created = Section.objects.get_or_create(
            tenant=school.tenant,
            school_class=school_class,
            name=section_name,
        )
        if created:
            summary["sections"]["created"] += 1
        else:
            summary["sections"]["skipped"] += 1
        section_map[(_normalize_key(school_class.name), _normalize_key(section_name))] = section

        if class_teacher_email:
            pending_class_teacher_links.append(
                {"email": class_teacher_email, "school_class": school_class, "section": section}
            )

    for index, row in enumerate(teacher_rows):
        full_name = str(row.get("teacher_name") or row.get("full_name") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        employee_id = str(row.get("employee_id") or "").strip()
        subjects_raw = str(row.get("subjects") or "").strip()
        class_teacher_of = str(row.get("class_teacher_of") or "").strip()

        if not full_name or not email:
            add_error("teachers", index, "teacher_name and email are required.")
            summary["teachers"]["skipped"] += 1
            continue

        existing_user = get_user_model().objects.filter(email__iexact=email).first()
        teacher = teacher_email_map.get(_normalize_key(email))
        if existing_user and teacher is None:
            add_error("teachers", index, f"Email already belongs to another account: {email}")
            summary["teachers"]["skipped"] += 1
            continue

        if teacher is None:
            generated_employee_id = employee_id or f"IMP-TCH-{index + 1:04d}"
            employee_id_candidate = generated_employee_id
            suffix = 1
            while Teacher.objects.filter(tenant=school.tenant, employee_id=employee_id_candidate).exists():
                suffix += 1
                employee_id_candidate = f"{generated_employee_id}-{suffix}"

            user = existing_user or get_user_model().objects.create_user(
                email=email,
                password=_make_import_password("Teacher", index + 1),
                role=UserRole.TEACHER,
                school=school,
            )
            teacher = Teacher.objects.create(
                tenant=school.tenant,
                user=user,
                employee_id=employee_id_candidate,
                full_name=full_name,
            )
            summary["teachers"]["created"] += 1
        else:
            if teacher.full_name != full_name:
                teacher.full_name = full_name
                teacher.save(update_fields=["full_name"])
                summary["teachers"]["updated"] += 1
            else:
                summary["teachers"]["skipped"] += 1

        teacher_email_map[_normalize_key(email)] = teacher

        assigned_section = None
        assigned_class = None
        if class_teacher_of:
            class_part, separator, section_part = class_teacher_of.partition("-")
            if separator and section_part:
                assigned_class, assigned_section = resolve_section(class_part, section_part)
            else:
                add_error("teachers", index, f"Invalid class_teacher_of format: {class_teacher_of}")

        if assigned_class and assigned_section:
            ClassTeacher.objects.get_or_create(
                tenant=school.tenant,
                teacher=teacher,
                school_class=assigned_class,
                section=assigned_section,
            )

        if assigned_section and subjects_raw:
            for subject_name in [item.strip() for item in subjects_raw.split(",") if item.strip()]:
                subject, _ = Subject.objects.get_or_create(
                    tenant=school.tenant,
                    name=subject_name,
                    defaults={"code": slugify(subject_name).replace("-", "").upper()[:20]},
                )
                TeachingAssignment.objects.get_or_create(
                    teacher=teacher,
                    section=assigned_section,
                    subject=subject,
                )

    for link in pending_class_teacher_links:
        teacher = teacher_email_map.get(_normalize_key(link["email"]))
        if teacher:
            ClassTeacher.objects.get_or_create(
                tenant=school.tenant,
                teacher=teacher,
                school_class=link["school_class"],
                section=link["section"],
            )

    for index, row in enumerate(student_rows):
        admission_number = str(row.get("admission_no") or row.get("admission_number") or "").strip()
        full_name = str(row.get("student_name") or row.get("full_name") or "").strip()
        class_identifier = row.get("class_code") or row.get("class_name")
        section_name = str(row.get("section_name") or "").strip()
        parent_phone = str(row.get("parent_phone") or "").strip()

        if not admission_number or not full_name or not class_identifier or not section_name:
            add_error("students", index, "admission_no, student_name, class_code, and section_name are required.")
            summary["students"]["skipped"] += 1
            continue

        school_class, section = resolve_section(class_identifier, section_name)
        if school_class is None:
            add_error("students", index, f"Unknown class reference: {class_identifier}")
            summary["students"]["skipped"] += 1
            continue
        if section is None:
            add_error("students", index, f"Unknown section reference: {section_name}")
            summary["students"]["skipped"] += 1
            continue

        student = Student.objects.filter(tenant=school.tenant, admission_number=admission_number).first()
        if student is None:
            Student.objects.create(
                tenant=school.tenant,
                admission_number=admission_number,
                full_name=full_name,
                parent_contact=parent_phone,
                school_class=school_class,
                section=section,
            )
            summary["students"]["created"] += 1
        else:
            updates = []
            if student.full_name != full_name:
                student.full_name = full_name
                updates.append("full_name")
            if parent_phone and student.parent_contact != parent_phone:
                student.parent_contact = parent_phone
                updates.append("parent_contact")
            if student.school_class_id != school_class.id:
                student.school_class = school_class
                updates.append("school_class")
            if student.section_id != section.id:
                student.section = section
                updates.append("section")
            if updates:
                student.save(update_fields=updates)
                summary["students"]["updated"] += 1
            else:
                summary["students"]["skipped"] += 1

    return {
        "detail": "Bulk import completed.",
        "school_uuid": str(school.uuid),
        "summary": summary,
        "errors": errors[:50],
    }


class SchoolListCreateAPI(APIView):

    def get(self, request):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        schools = School.objects.all().order_by("-created_at")
        serializer = SchoolSerializer(schools, many=True)
        return Response(serializer.data)

    def post(self, request):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        serializer = SchoolCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        creator = _resolve_tenant_creator(request)

        with transaction.atomic():
            contact_email = serializer.validated_data.get("contact_email") or getattr(creator, "email", "")
            tenant = Tenant.objects.create(
                creator=creator,
                name=serializer.validated_data["name"],
                type=TenantType.ORGANIZATION,
                billing_email=contact_email,
            )
            school = serializer.save(tenant=tenant)

        return Response(SchoolSerializer(school).data, status=status.HTTP_201_CREATED)


class SchoolDetailAPI(APIView):

    def get(self, request, school_uuid):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        school = get_object_or_404(School, uuid=school_uuid)
        return Response(SchoolSerializer(school).data)

    def patch(self, request, school_uuid):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        school = get_object_or_404(School, uuid=school_uuid)
        serializer = SchoolCreateSerializer(school, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        school = serializer.save()
        return Response(SchoolSerializer(school).data)


class SchoolBrandingLookupAPI(APIView):
    authentication_classes = ()
    permission_classes = ()

    def get(self, request, code):
        school = get_object_or_404(School, code__iexact=code, is_active=True)
        return Response(SchoolSerializer(school, context={"request": request}).data)


class CurrentSchoolAPI(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request):
        if not is_school_admin(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, "school", None)
        if not school:
            return Response({"detail": "School is not configured."}, status=status.HTTP_400_BAD_REQUEST)
        return Response(SchoolSerializer(school, context={"request": request}).data)

    def patch(self, request):
        if not is_school_admin(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, "school", None)
        if not school:
            return Response({"detail": "School is not configured."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SchoolCreateSerializer(school, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        school = serializer.save()
        return Response(SchoolSerializer(school, context={"request": request}).data)


class SchoolAdminCreateAPI(APIView):
    def post(self, request, school_uuid):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        school = get_object_or_404(School, uuid=school_uuid)
        serializer = SchoolAdminCreateSerializer(data=request.data, context={"school": school})
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user = serializer.save()
        return Response(
            {"uuid": str(user.uuid), "email": user.email, "role": user.role, "school_uuid": str(school.uuid)},
            status=status.HTTP_201_CREATED,
        )


class SchoolBulkImportAPI(APIView):
    def post(self, request, school_uuid):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        school = get_object_or_404(School, uuid=school_uuid)
        if not school.tenant:
            return Response({"detail": "School tenant is not configured."}, status=status.HTTP_400_BAD_REQUEST)

        payload = request.data if isinstance(request.data, dict) else {}
        return Response(_run_school_bulk_import(school, payload), status=status.HTTP_200_OK)


class CurrentSchoolBulkImportAPI(APIView):
    def post(self, request):
        if not is_school_admin(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, "school", None)
        if not school or not school.tenant:
            return Response({"detail": "School tenant is not configured."}, status=status.HTTP_400_BAD_REQUEST)
        payload = request.data if isinstance(request.data, dict) else {}
        return Response(_run_school_bulk_import(school, payload), status=status.HTTP_200_OK)


class CurrentSchoolImportPreviewAPI(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        if not is_school_admin(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        import_group = str(request.data.get("import_group") or "").strip().lower()
        uploaded_file = request.FILES.get("file")
        if import_group not in IMPORT_GROUP_FIELDS:
            return Response({"detail": "Invalid import group."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file:
            return Response({"detail": "Upload a file first."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            headers, rows = _parse_uploaded_table(uploaded_file)
        except Exception as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(
            {
                "headers": headers,
                "row_count": len(rows),
                "required_fields": IMPORT_GROUP_FIELDS[import_group],
            }
        )


class CurrentSchoolFileImportAPI(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        if not is_school_admin(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)
        school = getattr(request.user, "school", None)
        if not school or not school.tenant:
            return Response({"detail": "School tenant is not configured."}, status=status.HTTP_400_BAD_REQUEST)
        import_group = str(request.data.get("import_group") or "").strip().lower()
        uploaded_file = request.FILES.get("file")
        mapping_raw = request.data.get("mapping") or "{}"
        if import_group not in IMPORT_GROUP_FIELDS:
            return Response({"detail": "Invalid import group."}, status=status.HTTP_400_BAD_REQUEST)
        if not uploaded_file:
            return Response({"detail": "Upload a file first."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            mapping = json.loads(mapping_raw) if isinstance(mapping_raw, str) else mapping_raw
            headers, rows = _parse_uploaded_table(uploaded_file)
        except Exception as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        normalized_rows = _rows_with_mapping(headers, rows, mapping or {})
        payload = {key: [] for key in IMPORT_GROUP_FIELDS}
        payload[import_group] = normalized_rows
        return Response(_run_school_bulk_import(school, payload), status=status.HTTP_200_OK)


class SchoolSubscriptionAssignAPI(APIView):
    def post(self, request, school_uuid):
        if not _is_super_admin_or_open(request.user):
            return Response({"detail": "Not allowed"}, status=status.HTTP_403_FORBIDDEN)

        school = get_object_or_404(School, uuid=school_uuid)
        if not school.tenant:
            return Response({"detail": "School tenant is not configured."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SchoolSubscriptionAssignSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        plan_config = serializer.get_plan_config()
        price = Price.objects.get_by_plan(plan_config)
        if not price:
            return Response(
                {
                    "detail": "Subscription skipped because the plan is not initialized yet.",
                    "schedule_id": None,
                    "plan": plan_config.name,
                },
                status=status.HTTP_200_OK,
            )

        schedule = subscriptions.get_schedule(tenant=school.tenant)
        if schedule:
            current_phase = subscriptions.get_current_schedule_phase(schedule)
            current_phase["end_date"] = "now"
            next_phase = {"items": [{"price": price.id}]}
            schedule = subscriptions.update_schedule(
                schedule,
                phases=[current_phase, next_phase],
            )
        else:
            schedule = subscriptions.create_schedule(
                tenant=school.tenant,
                price=price,
            )

        return Response(
            {
                "detail": "Subscription updated.",
                "schedule_id": schedule.id,
                "plan": plan_config.name,
            }
        )
